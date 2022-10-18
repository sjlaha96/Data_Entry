from tkinter import *
from tkinter import font
from tkinter.ttk import Combobox
import tkinter as tk
from tkinter import messagebox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib
from pathlib import Path

root=Tk()
root.title("Data Fill Up")
root.geometry("700x500+300+200")
root.resizable(False,False)
root.config(bg="yellow")

file=pathlib.Path('Data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="FULL NAME"
    sheet['B1']="CONTACT"
    sheet['C1']="AGE"
    sheet['D1']="GENDER"
    sheet['E1']="ADDRESS"
    
    file.save('Data.xlsx')
    
def submit():
    name=nameValue.get()
    contact=contactValue.get()
    age=ageValue.get()
    gender=gender_combo.get()
    address=addressEntry.get(1.0,END)
    
    file=openpyxl.load_workbook('Data.xlsx')
    sheet=file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=name)
    sheet.cell(column=2,row=sheet.max_row,value=contact)
    sheet.cell(column=3,row=sheet.max_row,value=age)
    sheet.cell(column=4,row=sheet.max_row,value=gender)
    sheet.cell(column=5,row=sheet.max_row,value=address)
    
    file.save(r'Data.xlsx')
    
    messagebox.showinfo('SUCCESS','Details Added')
    nameValue.set("")
    contactValue.set("")
    ageValue.set("")
    addressEntry.delete(1.0,END)
    

def clear():
    nameValue.set("")
    contactValue.set("")
    ageValue.set("")
    addressEntry.delete(1.0,END)

icon_image=PhotoImage(file="data.png")
root.iconphoto(False,icon_image)

Label(root,text="PLEASE ENTER YOUR DETAILS",font=("arial",15),fg="black",bg="yellow").place(x=215,y=20)

Label(root,text="NAME",font=("arial",17),fg="black",bg="yellow").place(x=20,y=80)
Label(root,text="CONTACT",font=("arial",17),fg="black",bg="yellow").place(x=20,y=140)
Label(root,text="AGE",font=("arial",17),fg="black",bg="yellow").place(x=20,y=200)
Label(root,text="GENDER",font=("arial",17),fg="black",bg="yellow").place(x=20,y=260)
Label(root,text="ADDRESS",font=("arial",17),fg="black",bg="yellow").place(x=20,y=320)

nameValue=StringVar()
contactValue=StringVar()
ageValue=StringVar()

nameEntry=Entry(root,textvariable=nameValue,width=32,font=("arial",17),bd=2)
nameEntry.place(x=160,y=80)
contactEntry=Entry(root,textvariable=contactValue,width=32,font=("arial",17),bd=2)
contactEntry.place(x=160,y=140)
ageEntry=Entry(root,textvariable=ageValue,width=32,font=("arial",17),bd=2)
ageEntry.place(x=160,y=200)

gender_combo=Combobox(root,values=["Male","Female","Transgender","Others"],font=("arial",15),width=20)
gender_combo.place(x=160,y=260)
gender_combo.set("Select Gender")

addressEntry=Text(root,width=32,height=4,bd=4,font=("arial",17))
addressEntry.place(x=160,y=320)



Button(root,text="SUBMIT",fg="black",bg="grey",height=2,width=15,command=submit).place(x=160,y=450)
Button(root,text="CLEAR",fg="black",bg="grey",height=2,width=15,command=clear).place(x=300,y=450)
Button(root,text="EXIT",fg="black",bg="grey",height=2,width=15,command=lambda:root.destroy()).place(x=440,y=450)


root.mainloop()